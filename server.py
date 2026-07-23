#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import mimetypes
import os
import re
import socket
import sys
import tempfile
import threading
import unicodedata
import webbrowser
from difflib import SequenceMatcher
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

sys.path.insert(0, str(Path(__file__).resolve().parent / ".vendor"))

import qrcode
from openpyxl import load_workbook
from PIL import Image, ImageOps


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "Food panda LAB.xlsx"
IMAGE_DIR = ROOT / "image"
CACHE_DIR = ROOT / ".menu-cache"
ALLOWED_IMAGES = {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"}
STATIC_FILES = {
    "/": "index.html",
    "/index.html": "index.html",
    "/styles.css": "styles.css",
    "/app.js": "app.js",
}
STOP_WORDS = {"with"}
QR_PREFIX = "latte_lab_menu_qr_"

_menu_lock = threading.Lock()
_menu_cache: dict[str, object] = {"signature": None, "data": None}
_qr_lock = threading.Lock()
_qr_state: dict[str, object] = {"url": None, "path": None}


def log_event(message: str, error: bool = False) -> None:
    stream = sys.stderr if error else sys.stdout
    if stream is not None:
        print(message, file=stream, flush=True)
        return
    try:
        with (ROOT / ".server.log").open("a", encoding="utf-8") as log:
            log.write(f"{message}\n")
    except OSError:
        pass


def normalized_name(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return " ".join(re.findall(r"[a-z0-9]+", text))


def token_key(value: object) -> str:
    return " ".join(sorted(token for token in normalized_name(value).split() if token not in STOP_WORDS))


def image_files() -> list[Path]:
    if not IMAGE_DIR.exists():
        return []
    return sorted(
        (
            path
            for path in IMAGE_DIR.iterdir()
            if path.is_file()
            and path.suffix.casefold() in ALLOWED_IMAGES
            and not path.name.startswith(".")
            and not path.name.startswith(QR_PREFIX)
        ),
        key=lambda path: path.name.casefold(),
    )


def find_logo(files: list[Path]) -> Path | None:
    matches = [path for path in files if normalized_name(path.stem) == "logo"]
    return max(matches, key=lambda path: path.stat().st_mtime_ns) if matches else None


def find_image(name: str, files: list[Path]) -> Path | None:
    candidates = [path for path in files if normalized_name(path.stem) != "logo"]
    normalized = normalized_name(name)
    keyed = token_key(name)

    exact_matches = [path for path in candidates if normalized_name(path.stem) == normalized]
    if exact_matches:
        return max(exact_matches, key=lambda path: path.stat().st_mtime_ns)

    token_matches = [path for path in candidates if token_key(path.stem) == keyed]
    if token_matches:
        return max(token_matches, key=lambda path: path.stat().st_mtime_ns)

    best: tuple[float, Path | None] = (0.0, None)
    for path in candidates:
        ratio = SequenceMatcher(None, normalized, normalized_name(path.stem)).ratio()
        if ratio > best[0]:
            best = (ratio, path)
    return best[1] if best[0] >= 0.82 else None


def truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    return str(value or "").strip().casefold() in {"1", "yes", "true", "y"}


def number_or_none(value: object) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else number


def source_signature(files: list[Path]) -> str:
    parts = []
    if WORKBOOK_PATH.exists():
        stat = WORKBOOK_PATH.stat()
        parts.append(f"xlsx:{stat.st_mtime_ns}:{stat.st_size}")
    parts.extend(f"{path.name}:{path.stat().st_mtime_ns}:{path.stat().st_size}" for path in files)
    return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:16]


def media_url(path: Path | None, width: int) -> str | None:
    if path is None:
        return None
    return f"/media/{quote(path.name)}?w={width}"


def menu_url() -> str:
    return f"http://{local_ip()}:{os.environ.get('PORT', '8080')}/"


def ensure_qr_file(force: bool = False) -> tuple[Path, str]:
    url = menu_url()
    safe_address = re.sub(r"[^0-9A-Za-z]+", "-", urlparse(url).netloc).strip("-")
    target = IMAGE_DIR / f"{QR_PREFIX}{safe_address}.png"

    with _qr_lock:
        if not force and _qr_state["url"] == url and target.exists():
            return target, url

        IMAGE_DIR.mkdir(exist_ok=True)
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=14,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)
        image = qr.make_image(fill_color="#1c4028", back_color="#ffffff")

        with tempfile.NamedTemporaryFile(dir=IMAGE_DIR, prefix=".qr-", suffix=".png", delete=False) as temp:
            temp_path = Path(temp.name)
        try:
            image.save(temp_path, "PNG")
            for old_qr in IMAGE_DIR.glob(f"{QR_PREFIX}*.png"):
                old_qr.unlink(missing_ok=True)
            os.replace(temp_path, target)
        finally:
            temp_path.unlink(missing_ok=True)

        _qr_state.update(url=url, path=target)
        log_event(f"QR code updated: {target.name} -> {url}")
        return target, url


def read_menu() -> dict[str, object]:
    files = image_files()
    qr_path, current_menu_url = ensure_qr_file()
    signature = f"{source_signature(files)}-{qr_path.stem}"
    with _menu_lock:
        if _menu_cache["signature"] == signature:
            return _menu_cache["data"]  # type: ignore[return-value]

        workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (
                index
                for index, row in enumerate(rows[:12])
                if "Name" in row and "Type" in row and "Shop Price $" in row and "Shop Price ៛" in row
            ),
            None,
        )
        if header_index is None:
            raise ValueError("Could not find menu headers in Food panda LAB.xlsx")

        headers = [str(value or "").strip() for value in rows[header_index]]
        columns = {name: index for index, name in enumerate(headers)}
        logo = find_logo(files)
        items: list[dict[str, object]] = []
        categories: dict[str, int] = {}
        missing_images: list[str] = []

        for fallback_id, row in enumerate(rows[header_index + 1 :], start=1):
            name = str(row[columns["Name"]] or "").strip()
            if not name:
                continue

            category = str(row[columns["Type"]] or "Other").strip() or "Other"
            image_path = find_image(name, files)
            item = {
                "id": number_or_none(row[columns.get("No", -1)]) or fallback_id,
                "category": category,
                "en": name,
                "km": str(row[columns.get("Khmer", -1)] or "").strip(),
                "usd": number_or_none(row[columns["Shop Price $"]]),
                "khr": number_or_none(row[columns["Shop Price ៛"]]),
                "promotion": truthy(row[columns.get("Promotion", -1)]),
                "image": media_url(image_path, 760),
                "fullImage": media_url(image_path, 1800),
            }
            items.append(item)
            categories[category] = categories.get(category, 0) + 1
            if image_path is None:
                missing_images.append(name)

        data = {
            "version": signature,
            "logo": media_url(logo, 180),
            "items": items,
            "promotions": [item for item in items if item["promotion"]],
            "categories": [{"name": name, "count": count} for name, count in categories.items()],
            "missingImages": missing_images,
            "menuUrl": current_menu_url,
            "qrVersion": qr_path.stem,
        }
        _menu_cache.update(signature=signature, data=data)
        return data


def optimized_image(source: Path, width: int) -> Path:
    width = max(64, min(width, 2400))
    stat = source.stat()
    digest = hashlib.sha1(f"{source.name}:{stat.st_mtime_ns}:{stat.st_size}:{width}".encode()).hexdigest()[:16]
    target = CACHE_DIR / f"{digest}.webp"
    if target.exists():
        return target

    CACHE_DIR.mkdir(exist_ok=True)
    with Image.open(source) as image:
        image = ImageOps.exif_transpose(image).convert("RGB")
        image.thumbnail((width, width), Image.Resampling.LANCZOS)
        with tempfile.NamedTemporaryFile(dir=CACHE_DIR, suffix=".webp", delete=False) as temp:
            temp_path = Path(temp.name)
        try:
            image.save(temp_path, "WEBP", quality=86, method=6)
            os.replace(temp_path, target)
        finally:
            temp_path.unlink(missing_ok=True)
    return target


class MenuHandler(BaseHTTPRequestHandler):
    server_version = "LatteLabMenu/1.0"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/menu":
                self.serve_json(read_menu())
                return
            if parsed.path == "/qr.png":
                qr_path, _ = ensure_qr_file()
                self.serve_bytes(qr_path.read_bytes(), "image/png", "no-store")
                return
            if parsed.path.startswith("/media/"):
                self.serve_media(unquote(parsed.path.removeprefix("/media/")), parsed.query)
                return
            if parsed.path == "/favicon.ico":
                logo = find_logo(image_files())
                if logo:
                    self.serve_media(logo.name, "w=64")
                else:
                    self.send_error(HTTPStatus.NOT_FOUND)
                return
            if parsed.path in STATIC_FILES:
                self.serve_static(STATIC_FILES[parsed.path])
                return
            self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as error:
            log_event(f"Request error: {error}", error=True)
            self.send_error(HTTPStatus.INTERNAL_SERVER_ERROR)

    def serve_json(self, payload: object) -> None:
        body = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_bytes(self, body: bytes, content_type: str, cache_control: str) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", cache_control)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_static(self, filename: str) -> None:
        path = ROOT / filename
        body = path.read_bytes()
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", f"{content_type}; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_media(self, filename: str, query: str) -> None:
        safe_name = Path(filename).name
        if safe_name != filename:
            self.send_error(HTTPStatus.BAD_REQUEST)
            return
        source = IMAGE_DIR / safe_name
        if not source.is_file() or source.suffix.casefold() not in ALLOWED_IMAGES:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        params = parse_qs(query)
        try:
            width = int(params.get("w", ["1200"])[0])
        except ValueError:
            width = 1200
        target = optimized_image(source, width)
        body = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "image/webp")
        self.send_header("Cache-Control", "public, max-age=86400")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format_string: str, *args: object) -> None:
        log_event(f"{self.address_string()} - {format_string % args}")


def local_ip() -> str:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as connection:
            connection.connect(("8.8.8.8", 80))
            return connection.getsockname()[0]
    except OSError:
        pass

    try:
        addresses = socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET)
        for address in addresses:
            candidate = address[4][0]
            if not candidate.startswith(("127.", "169.254.")):
                return candidate
    except OSError:
        pass
    return "127.0.0.1"


def main() -> None:
    port = int(os.environ.get("PORT", "8080"))
    os.environ["PORT"] = str(port)
    ensure_qr_file(force=True)
    try:
        server = ThreadingHTTPServer(("0.0.0.0", port), MenuHandler)
    except OSError:
        if "--open-browser" in sys.argv:
            webbrowser.open(f"http://127.0.0.1:{port}/")
        log_event(f"Port {port} is already in use.", error=True)
        return
    (ROOT / ".server.pid").write_text(str(os.getpid()))
    (ROOT / ".server.port").write_text(str(port))
    url = menu_url()
    log_event(f"Latte Lab menu: {url}")
    if "--open-browser" in sys.argv:
        threading.Timer(0.7, lambda: webbrowser.open(url)).start()
    server.serve_forever()


if __name__ == "__main__":
    main()
